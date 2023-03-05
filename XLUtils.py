import openpyxl
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def getColumnCOunt(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum, column=columnnum).value

def writeData(file,sheetName,rounum,columnnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row=rounum, column=columnnum).value=data
    workbook.save(file)